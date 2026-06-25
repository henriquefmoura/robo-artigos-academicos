from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
import xml.etree.ElementTree as ET
from copy import copy
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup


BASE = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE / "Favaretto__Henrique_G04__Digital_Retail__25_04_2026_original.xlsx"
REPORT_DIR = BASE / "relatorios_robo"
CACHE_PATH = REPORT_DIR / "cache_fontes_abstract.json"
DEFAULT_REFERENCE_PATTERNS = ["*copilot*.xlsx", "*final*.xlsx", "*validado*.xlsx"]

SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": "Mozilla/5.0 abstract-cleaner-rpa/1.0 (mailto:research@example.com)",
        "Accept": "application/json,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
)

EMPTY_VALUES = {"", "x", "nan", "none", "null", "na", "n/a", "-", "--"}
BOOK_TYPES = {
    "book",
    "book-chapter",
    "book-part",
    "book-section",
    "book-series",
    "edited-book",
    "monograph",
    "reference-book",
    "reference-entry",
}
ARTICLE_TYPES = {
    "article",
    "journal-article",
    "proceedings-article",
    "preprint",
    "posted-content",
    "review",
}
SECTION_LABELS = [
    "abstract",
    "summary",
    "resumo",
    "purpose",
    "background",
    "objectives?",
    "aims?",
    "design/methodology/approach",
    "methodology",
    "methods?",
    "results?",
    "findings?",
    "conclusions?",
    "conclusion",
    "originality/value",
    "practical implications",
    "social implications",
    "theoretical implications",
    "limitations?",
]


@dataclass
class RowReport:
    row: int
    doi: str
    title: str
    before_status: str
    after_status: str
    source: str
    publication_type: str
    action: str
    note: str = ""


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty_abstract(value: Any) -> bool:
    return norm(value).lower() in EMPTY_VALUES


def looks_dirty(value: str) -> bool:
    return bool(re.search(r"<[^>]+>|&lt;/?|jats:|</jats|<jats", value, re.I))


def status_of(value: Any) -> str:
    text = norm(value)
    if is_empty_abstract(text):
        return "vazio"
    if looks_dirty(text):
        return "sujo"
    return "ok"


def collapse_spaces(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+([,.;:!?])", r"\1", text)
    return text.strip()


def remove_common_labels(text: str) -> str:
    labels = "|".join(SECTION_LABELS)
    text = re.sub(rf"^\s*(?:{labels})\s*[:.\-–—]?\s*", "", text, flags=re.I)
    text = re.sub(rf"(?<=[.!?])\s+(?:{labels})\s*[:.\-–—]\s*", " ", text, flags=re.I)
    text = re.sub(rf"\b(?:{labels})\s*[:]\s*", "", text, flags=re.I)
    return text


def clean_abstract(raw: Any) -> str:
    text = norm(raw)
    if is_empty_abstract(text):
        return ""

    text = html.unescape(text)
    text = re.sub(r"^\s*<!\[CDATA\[|\]\]>\s*$", "", text, flags=re.I)

    if looks_dirty(text):
        soup = BeautifulSoup(f"<root>{text}</root>", "xml")
        if soup.find() is None:
            soup = BeautifulSoup(text, "html.parser")
        for tag in soup.find_all(["title", "jats:title"]):
            tag.decompose()
        for tag in soup.find_all(True):
            name = (tag.name or "").lower()
            if name.endswith(":title"):
                tag.decompose()
        text = soup.get_text(" ", strip=True)

    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = collapse_spaces(text)
    text = remove_common_labels(text)
    text = collapse_spaces(text)
    text = re.sub(r"^\W*(abstract|summary|resumo)\W*", "", text, flags=re.I)
    return collapse_spaces(text)


def abstract_from_openalex_inverted(index: dict[str, list[int]] | None) -> str:
    if not index:
        return ""
    words: list[str | None] = [None] * (max(pos for positions in index.values() for pos in positions) + 1)
    for word, positions in index.items():
        for pos in positions:
            words[pos] = word
    return " ".join(word for word in words if word)


def get_json(url: str, timeout: int = 25) -> dict[str, Any] | None:
    try:
        response = SESSION.get(url, timeout=timeout)
        if response.status_code == 200:
            return response.json()
    except Exception:
        return None
    return None


def load_cache() -> dict[str, dict[str, str]]:
    if CACHE_PATH.exists():
        try:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache: dict[str, dict[str, str]]) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", norm(value).lower())


def valid_reference_abstract(value: Any) -> bool:
    text = norm(value)
    return bool(text) and text.lower() not in EMPTY_VALUES and not looks_dirty(text)


def discover_reference_workbook(input_path: Path) -> Path | None:
    candidates: list[Path] = []
    for pattern in DEFAULT_REFERENCE_PATTERNS:
        candidates.extend(input_path.parent.glob(pattern))
    candidates = [
        p
        for p in candidates
        if p.is_file()
        and p.resolve() != input_path.resolve()
        and not p.name.startswith("~$")
        and not p.name.startswith("_tmp_")
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def copy_locked_workbook(path: Path) -> Path:
    temp = path.with_name(f"_tmp_robo_ref_{path.stem}.xlsx")
    temp.write_bytes(path.read_bytes())
    return temp


def load_reference_abstracts(reference_path: Path | None) -> dict[str, dict[str, str]]:
    if not reference_path:
        return {}
    source_path = reference_path
    temp_path: Path | None = None
    try:
        try:
            wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        except PermissionError:
            temp_path = copy_locked_workbook(reference_path)
            source_path = temp_path
            wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

        refs: dict[str, dict[str, str]] = {}
        for ws in wb.worksheets:
            headers = [norm(cell.value).lower() for cell in ws[1]]
            if "abstract" not in headers:
                continue
            abstract_col = find_header(ws, "abstract")
            doi_col = find_header(ws, "doi") if "doi" in headers else None
            title_col = find_header(ws, "titulo") if "titulo" in headers else None
            for row in range(2, ws.max_row + 1):
                abstract = clean_abstract(ws.cell(row, abstract_col).value)
                if not valid_reference_abstract(abstract):
                    continue
                doi = normalize_key(ws.cell(row, doi_col).value) if doi_col else ""
                title = normalize_key(ws.cell(row, title_col).value) if title_col else ""
                payload = {
                    "abstract": abstract,
                    "classification": "livro" if abstract.lower() == "(livro)" else "artigo",
                    "source": f"referencia:{reference_path.name}",
                    "publication_type": "reference-workbook",
                }
                if doi:
                    refs[f"doi:{doi}"] = payload
                if title:
                    refs[f"title:{title}"] = payload
        return refs
    finally:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def fetch_reference(
    doi: str,
    title: str,
    references: dict[str, dict[str, str]],
) -> tuple[str, str, str, str]:
    for key in [f"doi:{normalize_key(doi)}", f"title:{normalize_key(title)}"]:
        if key in references:
            item = references[key]
            return (
                item.get("abstract", ""),
                item.get("classification", "artigo"),
                item.get("source", "referencia"),
                item.get("publication_type", "reference-workbook"),
            )
    return "", "", "", ""


def fetch_crossref(doi: str, title: str) -> tuple[str, str, str]:
    if doi:
        url = f"https://api.crossref.org/works/{quote_plus(doi)}"
    else:
        url = f"https://api.crossref.org/works?query.title={quote_plus(title)}&rows=1"
    data = get_json(url)
    if not data:
        return "", "", ""
    message = data.get("message", {})
    item = message if doi else (message.get("items") or [{}])[0]
    pub_type = norm(item.get("type")).lower()
    abstract = clean_abstract(item.get("abstract", ""))
    return abstract, pub_type, "crossref"


def fetch_openalex(doi: str, title: str) -> tuple[str, str, str]:
    if doi:
        url = f"https://api.openalex.org/works/https://doi.org/{quote_plus(doi)}"
    else:
        url = f"https://api.openalex.org/works?search={quote_plus(title)}&per-page=1"
    data = get_json(url)
    if not data:
        return "", "", ""
    item = data if doi else (data.get("results") or [{}])[0]
    pub_type = norm(item.get("type") or item.get("type_crossref")).lower()
    abstract = clean_abstract(abstract_from_openalex_inverted(item.get("abstract_inverted_index")))
    return abstract, pub_type, "openalex"


def fetch_semantic_scholar(doi: str, title: str) -> tuple[str, str, str]:
    if doi:
        url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{quote_plus(doi)}?fields=title,abstract,publicationTypes"
    else:
        url = f"https://api.semanticscholar.org/graph/v1/paper/search?query={quote_plus(title)}&limit=1&fields=title,abstract,publicationTypes"
    data = get_json(url)
    if not data:
        return "", "", ""
    item = data if doi else (data.get("data") or [{}])[0]
    pub_types = item.get("publicationTypes") or []
    pub_type = ",".join(str(t).lower() for t in pub_types)
    abstract = clean_abstract(item.get("abstract", ""))
    return abstract, pub_type, "semantic_scholar"


def fetch_europe_pmc(doi: str, title: str) -> tuple[str, str, str]:
    if doi:
        query = f'DOI:"{doi}"'
    elif title:
        query = f'TITLE:"{title}"'
    else:
        return "", "", ""
    url = (
        "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
        f"?query={quote_plus(query)}&format=json&pageSize=1"
    )
    data = get_json(url)
    if not data:
        return "", "", ""
    item = ((data.get("resultList") or {}).get("result") or [{}])[0]
    pub_type = norm(item.get("pubType") or item.get("source")).lower()
    abstract = clean_abstract(item.get("abstractText", ""))
    return abstract, pub_type, "europe_pmc"


def fetch_doaj(doi: str, title: str) -> tuple[str, str, str]:
    query = doi or title
    if not query:
        return "", "", ""
    url = f"https://doaj.org/api/search/articles/{quote_plus(query)}?pageSize=1"
    data = get_json(url)
    if not data:
        return "", "", ""
    result = (data.get("results") or [{}])[0]
    bibjson = result.get("bibjson") or {}
    abstract = clean_abstract(bibjson.get("abstract", ""))
    return abstract, "journal-article", "doaj"


def fetch_arxiv(doi: str, title: str) -> tuple[str, str, str]:
    if doi:
        search = f"doi:{doi}"
    elif title:
        search = f'ti:"{title}"'
    else:
        return "", "", ""
    url = f"https://export.arxiv.org/api/query?search_query={quote_plus(search)}&max_results=1"
    try:
        response = SESSION.get(url, timeout=25)
        if response.status_code != 200:
            return "", "", ""
        root = ET.fromstring(response.text)
        ns = {"atom": "http://www.w3.org/2005/Atom"}
        entry = root.find("atom:entry", ns)
        if entry is None:
            return "", "", ""
        summary = entry.findtext("atom:summary", default="", namespaces=ns)
        return clean_abstract(summary), "preprint", "arxiv"
    except Exception:
        return "", "", ""


def fetch_doi_landing_meta(doi: str) -> tuple[str, str, str]:
    if not doi:
        return "", "", ""
    try:
        response = SESSION.get(f"https://doi.org/{doi}", timeout=30, allow_redirects=True)
        if response.status_code >= 400 or "html" not in response.headers.get("content-type", "").lower():
            return "", "", ""
        soup = BeautifulSoup(response.text, "html.parser")
        pub_type = ""
        for key in ["citation_abstract", "dc.description", "description", "og:description"]:
            attrs = {"name": key} if key != "og:description" else {"property": key}
            meta = soup.find("meta", attrs=attrs)
            if meta and meta.get("content"):
                return clean_abstract(meta["content"]), pub_type, "doi_landing_meta"
        for meta in soup.find_all("meta"):
            name = (meta.get("name") or meta.get("property") or "").lower()
            if "type" in name and meta.get("content"):
                pub_type = meta["content"].lower()
    except Exception:
        return "", "", ""
    return "", "", ""


def is_probable_abstract(text: str) -> bool:
    text = clean_abstract(text)
    lowered = text.lower()
    if len(text) < 90 or len(text) > 6000:
        return False
    bad_phrases = [
        "no abstract is available",
        "abstract is not available",
        "downloadable with restrictions",
        "semantic scholar extracted view",
        "please complete the following challenge",
        "enable javascript",
        "just a moment",
        "cookies to continue",
    ]
    if any(phrase in lowered for phrase in bad_phrases):
        return False
    research_markers = [
        "this study",
        "this paper",
        "this article",
        "we examine",
        "we investigate",
        "we explore",
        "we analyze",
        "we analyse",
        "findings",
        "results",
        "purpose",
        "method",
        "data",
        "research",
    ]
    return any(marker in lowered for marker in research_markers)


def extract_abstract_from_html(html_text: str) -> str:
    soup = BeautifulSoup(html_text, "html.parser")
    for tag in soup(["script", "style", "nav", "header", "footer", "aside", "form"]):
        tag.decompose()

    meta_keys = [
        "citation_abstract",
        "dc.description",
        "description",
        "og:description",
        "twitter:description",
    ]
    for key in meta_keys:
        attrs = {"name": key} if key not in {"og:description", "twitter:description"} else {"property": key}
        meta = soup.find("meta", attrs=attrs)
        if meta and meta.get("content"):
            candidate = clean_abstract(meta["content"])
            if is_probable_abstract(candidate):
                return candidate

    header_pattern = re.compile(r"^(abstract|summary|resumo)$", re.I)
    for header in soup.find_all(["h1", "h2", "h3", "h4", "h5", "dt", "strong", "b"]):
        if not header_pattern.match(header.get_text(" ", strip=True)):
            continue
        chunks: list[str] = []
        for sibling in header.find_all_next():
            if sibling.name in {"h1", "h2", "h3", "h4", "h5", "dt"} and sibling is not header:
                break
            text = sibling.get_text(" ", strip=True)
            if text:
                chunks.append(text)
            candidate = clean_abstract(" ".join(chunks))
            if is_probable_abstract(candidate):
                return candidate

    text = soup.get_text("\n", strip=True)
    match = re.search(
        r"\bAbstract\b\s*(.{120,5000}?)(?:\n\s*(?:Keywords?|Introduction|Suggested Citation|References|Cited by|Fingerprint)\b|$)",
        text,
        flags=re.I | re.S,
    )
    if match:
        candidate = clean_abstract(match.group(1))
        if is_probable_abstract(candidate):
            return candidate
    return ""


def duckduckgo_result_urls(query: str, limit: int = 5) -> list[str]:
    urls: list[str] = []
    try:
        response = SESSION.get(
            f"https://duckduckgo.com/html/?q={quote_plus(query)}",
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        if response.status_code not in {200, 202}:
            return urls
        if "Please complete the following challenge" in response.text:
            return urls
        soup = BeautifulSoup(response.text, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("//"):
                href = "https:" + href
            if "uddg=" in href:
                parsed = urlparse(href)
                target = parse_qs(parsed.query).get("uddg", [""])[0]
            elif href.startswith("http"):
                target = href
            else:
                continue
            if target and target not in urls:
                urls.append(target)
            if len(urls) >= limit:
                break
    except Exception:
        return urls
    return urls


def fetch_web_search_abstract(doi: str, title: str) -> tuple[str, str, str]:
    if not title:
        return "", "", ""
    network_terms = [
        "network",
        "networks",
        "neural network",
        "neural networks",
        "artificial neural network",
        "deep neural network",
        "convolutional neural network",
        "recurrent neural network",
        "graph neural network",
        "bayesian network",
        "social network",
        "complex network",
        "supply chain network",
        "retail network",
    ]
    queries = [
        f'"{title}" abstract',
        f'"{title}"',
    ]
    if doi:
        queries.insert(0, f'"{doi}" abstract')
    title_l = title.lower()
    if any(term in title_l for term in ["network", "neural", "graph", "bayesian"]):
        queries.extend(f'"{title}" "{term}" abstract' for term in network_terms)
    queries.extend(
        f'site:{domain} "{title}" abstract'
        for domain in [
            "openalex.org",
            "semanticscholar.org",
            "crossref.org",
            "europepmc.org",
            "doaj.org",
            "arxiv.org",
            "ssrn.com",
            "researchgate.net",
            "sciencedirect.com",
            "springer.com",
            "tandfonline.com",
            "emerald.com",
            "wiley.com",
            "sagepub.com",
            "mdpi.com",
            "frontiersin.org",
            "plos.org",
            "nature.com",
            "ieee.org",
            "acm.org",
        ]
    )

    blocked_domains = {"google.com", "bing.com", "duckduckgo.com"}
    preferred_domains = [
        "openalex.org",
        "crossref.org",
        "europepmc.org",
        "doaj.org",
        "sciencedirect.com",
        "inderscience.com",
        "ideas.repec.org",
        "econpapers.repec.org",
        "semanticscholar.org",
        "researchgate.net",
        "arxiv.org",
        "ssrn.com",
        "springer.com",
        "tandfonline.com",
        "emerald.com",
        "wiley.com",
        "sagepub.com",
        "mdpi.com",
        "frontiersin.org",
        "plos.org",
        "nature.com",
        "ieee.org",
        "acm.org",
        "jstor.org",
        "cambridge.org",
        "oup.com",
        "degruyter.com",
        "informs.org",
        "aisel.aisnet.org",
        "proquest.com",
        ".edu",
        ".ac.",
    ]

    seen: set[str] = set()
    for query in queries:
        urls = duckduckgo_result_urls(query, limit=10)
        urls.sort(key=lambda u: 0 if any(domain in urlparse(u).netloc.lower() for domain in preferred_domains) else 1)
        for url in urls[:8]:
            if url in seen:
                continue
            seen.add(url)
            domain = urlparse(url).netloc.lower()
            if any(domain.endswith(bad) for bad in blocked_domains):
                continue
            try:
                response = SESSION.get(url, timeout=30, allow_redirects=True)
                if response.status_code >= 400:
                    continue
                ctype = response.headers.get("content-type", "").lower()
                if "html" not in ctype and "xml" not in ctype and response.text[:100].lstrip()[:1] != "<":
                    continue
                abstract = extract_abstract_from_html(response.text)
                if abstract:
                    return abstract, "artigo", f"web:{domain}"
            except Exception:
                continue
            time.sleep(0.25)
    return "", "", ""


def classify_type(*types: str) -> str:
    joined = ",".join(t for t in types if t).lower()
    if any(book_type in joined for book_type in BOOK_TYPES):
        return "livro"
    if any(article_type in joined for article_type in ARTICLE_TYPES):
        return "artigo"
    return joined or "desconhecido"


def classify_title(title: str) -> str:
    title_l = title.lower().strip()
    if re.search(r"\bbook review\b|\bresenha\b", title_l):
        return "livro"
    if re.search(r"\bisbn\b|\bpaperback\b|\bhardcover\b|\bebook\b|\bpp\.\s*\d+", title_l):
        return "livro"
    return ""


def fetch_missing_abstract(
    doi: str,
    title: str,
    wait: float,
    cache: dict[str, dict[str, str]],
    references: dict[str, dict[str, str]] | None = None,
) -> tuple[str, str, str, str]:
    title_classification = classify_title(title)
    if title_classification == "livro":
        return "(livro)", "livro", "titulo", "book-review"

    cache_key = (doi or title).lower().strip()
    if cache_key in cache:
        cached = cache[cache_key]
        cached_abstract = cached.get("abstract", "")
        cached_classification = cached.get("classification", "")
        if cached_abstract or cached_classification == "livro":
            return (
                cached_abstract,
                cached_classification,
                cached.get("source", "cache"),
                cached.get("publication_type", ""),
            )

    found_types: list[str] = []
    for fetcher in [fetch_crossref, fetch_openalex, fetch_semantic_scholar, fetch_europe_pmc, fetch_doaj, fetch_arxiv]:
        abstract, pub_type, source = fetcher(doi, title)
        if pub_type:
            found_types.append(pub_type)
        classification = classify_type(*found_types)
        if classification == "livro":
            cache[cache_key] = {
                "abstract": "(livro)",
                "classification": "livro",
                "source": source,
                "publication_type": pub_type,
            }
            return "(livro)", "livro", source, pub_type
        if abstract and len(abstract) >= 80:
            cache[cache_key] = {
                "abstract": abstract,
                "classification": classification,
                "source": source,
                "publication_type": pub_type,
            }
            return abstract, classification, source, pub_type
        time.sleep(wait)

    abstract, pub_type, source = fetch_doi_landing_meta(doi)
    if pub_type:
        found_types.append(pub_type)
    classification = classify_type(*found_types)
    if classification == "livro":
        cache[cache_key] = {
            "abstract": "(livro)",
            "classification": "livro",
            "source": source,
            "publication_type": pub_type,
        }
        return "(livro)", "livro", source, pub_type
    if abstract and len(abstract) >= 80:
        cache[cache_key] = {
            "abstract": abstract,
            "classification": classification,
            "source": source,
            "publication_type": pub_type,
        }
        return abstract, classification, source, pub_type

    abstract, web_classification, web_source = fetch_web_search_abstract(doi, title)
    if abstract:
        classification = web_classification or classification
        cache[cache_key] = {
            "abstract": abstract,
            "classification": classification,
            "source": web_source,
            "publication_type": "web-search",
        }
        return abstract, classification, web_source, "web-search"

    if references:
        abstract, ref_classification, source, pub_type = fetch_reference(doi, title, references)
        if abstract:
            classification = ref_classification or classification
            cache[cache_key] = {
                "abstract": abstract,
                "classification": classification,
                "source": source,
                "publication_type": pub_type,
            }
            return abstract, classification, source, pub_type

    cache[cache_key] = {
        "abstract": "",
        "classification": classification,
        "source": "",
        "publication_type": ",".join(found_types),
    }
    return "", classification, "", ",".join(found_types)


def find_header(ws: openpyxl.worksheet.worksheet.Worksheet, wanted: str) -> int:
    wanted = wanted.lower()
    for cell in ws[1]:
        if norm(cell.value).lower() == wanted:
            return cell.column
    raise ValueError(f"Coluna obrigatoria nao encontrada: {wanted}")


def preserve_cell_style(src: openpyxl.cell.cell.Cell, dst: openpyxl.cell.cell.Cell) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def process_workbook(
    input_path: Path,
    output_path: Path,
    in_place: bool,
    wait: float,
    reference_path: Path | None,
    drop_empty_abstracts: bool = False,
    skip_missing_fetch: bool = False,
) -> list[RowReport]:
    wb = openpyxl.load_workbook(input_path)
    reports: list[RowReport] = []
    cache = load_cache()
    references = load_reference_abstracts(reference_path)
    if reference_path and references:
        print(f"Referencia carregada: {reference_path} ({len(references)} chaves)", flush=True)

    for ws in wb.worksheets:
        headers = [norm(cell.value).lower() for cell in ws[1]]
        if "abstract" not in headers:
            continue
        abstract_col = find_header(ws, "abstract")
        doi_col = find_header(ws, "doi") if "doi" in headers else None
        title_col = find_header(ws, "titulo") if "titulo" in headers else None

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, abstract_col)
            original = cell.value
            doi = norm(ws.cell(row, doi_col).value) if doi_col else ""
            title = norm(ws.cell(row, title_col).value) if title_col else ""
            before = status_of(original)
            action = "sem_alteracao"
            source = "planilha"
            pub_type = ""
            note = ""
            new_value = norm(original)

            if before == "sujo":
                cleaned = clean_abstract(original)
                if cleaned and cleaned != norm(original):
                    new_value = cleaned
                    action = "limpo"
            elif before == "vazio":
                if skip_missing_fetch:
                    action = "nao_encontrado"
                    note = "Busca externa de abstract pulada no modo rapido."
                else:
                    print(f"Linha {row}: buscando abstract ausente - {doi or title[:60]}", flush=True)
                    fetched, classification, source, raw_type = fetch_missing_abstract(doi, title, wait, cache, references)
                    save_cache(cache)
                    pub_type = raw_type or classification
                    if classification == "livro":
                        new_value = "(livro)"
                        action = "marcado_livro"
                    elif fetched:
                        new_value = fetched
                        action = "preenchido"
                    else:
                        action = "nao_encontrado"
                        note = "Resumo nao localizado nas fontes automaticas consultadas."

            if action in {"limpo", "preenchido", "marcado_livro"}:
                old_style = copy(cell._style)
                cell.value = new_value
                cell._style = old_style

            reports.append(
                RowReport(
                    row=row,
                    doi=doi,
                    title=title,
                    before_status=before,
                    after_status=status_of(cell.value),
                    source=source,
                    publication_type=pub_type,
                    action=action,
                    note=note,
                )
            )

        if drop_empty_abstracts:
            removed = 0
            for row in range(ws.max_row, 1, -1):
                abstract = norm(ws.cell(row, abstract_col).value)
                if is_empty_abstract(abstract):
                    doi = norm(ws.cell(row, doi_col).value) if doi_col else ""
                    title = norm(ws.cell(row, title_col).value) if title_col else ""
                    ws.delete_rows(row, 1)
                    removed += 1
                    reports.append(
                        RowReport(
                            row=row,
                            doi=doi,
                            title=title,
                            before_status="vazio",
                            after_status="removido",
                            source="filtro_final",
                            publication_type="",
                            action="removido_sem_abstract",
                            note="Linha removida porque nenhuma fonte automatica retornou abstract valido.",
                        )
                    )
            if removed:
                print(f"Linhas removidas sem abstract valido: {removed}", flush=True)

    target = input_path if in_place else output_path
    wb.save(target)
    save_cache(cache)
    return reports


def validation_counts(path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    blank_or_x = 0
    dirty_tags = 0
    book_marks = 0
    for ws in wb.worksheets:
        headers = [norm(cell.value).lower() for cell in ws[1]]
        if "abstract" not in headers:
            continue
        abstract_col = find_header(ws, "abstract")
        for row in range(2, ws.max_row + 1):
            text = norm(ws.cell(row, abstract_col).value)
            if is_empty_abstract(text):
                blank_or_x += 1
            if looks_dirty(text):
                dirty_tags += 1
            if text.lower() == "(livro)":
                book_marks += 1
    return {"blank_or_x": blank_or_x, "dirty_tags": dirty_tags, "book_marks": book_marks}


def write_reports(reports: list[RowReport], stem: str) -> tuple[Path, Path]:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = REPORT_DIR / f"{stem}_relatorio.json"
    csv_path = REPORT_DIR / f"{stem}_relatorio.csv"
    json_path.write_text(json.dumps([asdict(r) for r in reports], ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(reports[0]).keys()) if reports else [])
        if reports:
            writer.writeheader()
            writer.writerows(asdict(r) for r in reports)
    return json_path, csv_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Limpa a coluna abstract de planilhas XLSX e preenche resumos ausentes por fontes academicas.",
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Planilha XLSX de entrada.")
    parser.add_argument("--output", type=Path, default=None, help="Planilha XLSX de saida. Padrao: *_limpo.xlsx")
    parser.add_argument("--in-place", action="store_true", help="Sobrescreve a planilha de entrada.")
    parser.add_argument("--wait", type=float, default=1.0, help="Pausa entre consultas externas, em segundos.")
    parser.add_argument("--reference", type=Path, default=None, help="Planilha final/validada para usar como banco de abstracts por DOI/titulo.")
    parser.add_argument("--no-auto-reference", action="store_true", help="Nao procurar automaticamente arquivos *copilot*, *final* ou *validado* na pasta.")
    parser.add_argument("--drop-empty-abstracts", action="store_true", help="Remove linhas que continuarem sem abstract apos a busca.")
    parser.add_argument("--skip-missing-fetch", action="store_true", help="Nao buscar abstracts ausentes; apenas limpa existentes e aplica filtros finais.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve() if args.output else input_path.with_name(f"{input_path.stem}_limpo.xlsx")
    reference_path = args.reference.resolve() if args.reference else None
    if reference_path is None and not args.no_auto_reference:
        reference_path = discover_reference_workbook(input_path)
    reports = process_workbook(
        input_path,
        output_path,
        args.in_place,
        args.wait,
        reference_path,
        args.drop_empty_abstracts,
        args.skip_missing_fetch,
    )
    json_path, csv_path = write_reports(reports, output_path.stem)

    counts: dict[str, int] = {}
    for item in reports:
        counts[item.action] = counts.get(item.action, 0) + 1

    print(f"Planilha gerada: {input_path if args.in_place else output_path}")
    print(f"Relatorio JSON: {json_path}")
    print(f"Relatorio CSV: {csv_path}")
    print("Acoes:", counts)
    validation = validation_counts(input_path if args.in_place else output_path)
    print("Validacao:", validation)


if __name__ == "__main__":
    main()
