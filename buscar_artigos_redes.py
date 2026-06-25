from __future__ import annotations

import argparse
import html
import re
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus

import openpyxl
import requests
from bs4 import BeautifulSoup


BASE = Path(__file__).resolve().parent
DEFAULT_OUTPUT = BASE / "Corpus_busca_redes_retail.xlsx"
TEMPLATE_HEADERS = [
    "DOI",
    "titulo",
    "publicacao",
    "ano",
    "ajg_2024",
    "volume",
    "numero",
    "paginas",
    "autores",
    "abstract",
    "ajg_categ",
    "sco_quartil",
    "sco_publisher",
    "sco_pais",
    "sco_regiao",
    "consumer",
    "market",
    "secundario",
    "customer",
    "behavior",
    "retail",
    "centric",
    "shopper",
    "tem_S",
    "sco_categorias",
    "sco_areas",
    "Classificacao",
]

SESSION = requests.Session()
SESSION.mount("https://", requests.adapters.HTTPAdapter(pool_connections=32, pool_maxsize=32))
SESSION.mount("http://", requests.adapters.HTTPAdapter(pool_connections=32, pool_maxsize=32))
SESSION.headers.update(
    {
        "User-Agent": "Mozilla/5.0 article-network-search/1.0 (mailto:research@example.com)",
        "Accept": "application/json,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
)

NETWORK_TERMS = {
    "redes_neurais": [
        "neural network",
        "neural networks",
        "artificial neural network",
        "deep neural network",
        "convolutional neural network",
        "recurrent neural network",
        "LSTM network",
        "graph neural network",
        "transformer neural network",
    ],
    "redes_grafos_e_complexas": [
        "network analysis",
        "graph theory",
        "complex network",
        "small-world network",
        "scale-free network",
        "multilayer network",
        "network science",
    ],
    "redes_sociais_e_consumidor": [
        "social network",
        "online social network",
        "customer network",
        "consumer network",
        "word-of-mouth network",
        "influencer network",
    ],
    "redes_bayesianas_e_semanticas": [
        "bayesian network",
        "semantic network",
        "knowledge graph",
        "knowledge network",
        "causal network",
    ],
    "redes_operacoes_e_ecossistemas": [
        "supply chain network",
        "logistics network",
        "retail network",
        "distribution network",
        "platform network",
        "ecosystem network",
        "IoT network",
        "sensor network",
    ],
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def clean_abstract(raw: Any) -> str:
    text = norm(raw)
    if not text:
        return ""
    text = html.unescape(text)
    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)
    text = re.sub(r"^\W*(abstract|summary|resumo)\W*", "", text, flags=re.I)
    return norm(text)


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", norm(value).lower()).strip()


def abstract_from_openalex(index: dict[str, list[int]] | None) -> str:
    if not index:
        return ""
    size = max(pos for positions in index.values() for pos in positions) + 1
    words: list[str] = [""] * size
    for word, positions in index.items():
        for pos in positions:
            words[pos] = word
    return clean_abstract(" ".join(words))


def get_json(url: str, timeout: int = 30) -> dict[str, Any] | None:
    try:
        response = SESSION.get(url, timeout=timeout)
        if response.status_code == 200:
            return response.json()
    except Exception:
        return None
    return None


def doi_from_openalex(value: str) -> str:
    return norm(value).removeprefix("https://doi.org/")


def make_record(
    *,
    title: str,
    abstract: str = "",
    year: Any = "",
    journal: str = "",
    doi: str = "",
    url: str = "",
    authors: str = "",
    publication_type: str = "",
    source: str,
    network_group: str,
    network_term: str,
    search_query: str,
    citations: Any = "",
) -> dict[str, Any]:
    text = f"{title} {abstract} {search_query}".lower()
    record = {header: "" for header in TEMPLATE_HEADERS}
    record.update(
        {
            "DOI": norm(doi),
            "titulo": norm(title),
            "publicacao": norm(journal),
            "ano": year or "",
            "ajg_2024": "x",
            "volume": "x",
            "numero": "x",
            "paginas": "x",
            "autores": norm(authors),
            "abstract": clean_abstract(abstract),
            "ajg_categ": "x",
            "sco_quartil": "x",
            "sco_publisher": "x",
            "sco_pais": "x",
            "sco_regiao": "x",
            "consumer": marker(text, ["consumer", "consumidor"]),
            "market": marker(text, ["market", "mercado"]),
            "secundario": "",
            "customer": marker(text, ["customer", "cliente"]),
            "behavior": marker(text, ["behavior", "behaviour", "comportamento"]),
            "retail": marker(text, ["retail", "varejo"]),
            "centric": marker(text, ["centric", "centrado"]),
            "shopper": marker(text, ["shopper"]),
            "tem_S": marker(text, ["network", "neural", "graph", "retail", "digital", "customer", "consumer"]),
            "sco_categorias": "x",
            "sco_areas": "x",
            "Classificacao": "SIM" if clean_abstract(abstract) else "NAO",
            "_tipo_publicacao": norm(publication_type),
            "_citacoes": citations or "",
            "_fonte_busca": source,
            "_grupo_rede": network_group,
            "_termo_rede": network_term,
            "_consulta": search_query,
            "_url": norm(url),
        }
    )
    return record


def marker(text: str, terms: list[str]) -> str:
    return "S" if any(term in text for term in terms) else "N"


def merge_record(current: dict[str, Any] | None, incoming: dict[str, Any]) -> dict[str, Any]:
    if current is None:
        return incoming
    merged = current.copy()
    for key, value in incoming.items():
        if key.startswith("_"):
            if value and norm(value) not in norm(merged.get(key, "")):
                merged[key] = "; ".join(part for part in [norm(merged.get(key, "")), norm(value)] if part)
            continue
        if key == "abstract":
            if len(norm(value)) > len(norm(merged.get(key, ""))):
                merged[key] = value
            continue
        if key in {"consumer", "market", "customer", "behavior", "retail", "centric", "shopper", "tem_S"}:
            merged[key] = "S" if "S" in {norm(merged.get(key)), norm(value)} else "N"
            continue
        if is_blank_for_merge(merged.get(key)) and not is_blank_for_merge(value):
            merged[key] = value
    merged["Classificacao"] = "SIM" if norm(merged.get("abstract")) else "NAO"
    return merged


def is_blank_for_merge(value: Any) -> bool:
    return norm(value).lower() in {"", "x", "none", "nan", "null", "n/a", "na"}


def search_openalex(query: str, limit: int, network_group: str, network_term: str) -> list[dict[str, Any]]:
    url = f"https://api.openalex.org/works?search={quote_plus(query)}&per-page={limit}"
    data = get_json(url)
    if not data:
        return []
    records = []
    for item in data.get("results") or []:
        authorships = item.get("authorships") or []
        authors = "; ".join(norm((a.get("author") or {}).get("display_name")) for a in authorships if a.get("author"))
        source = item.get("primary_location") or {}
        source_info = source.get("source") or {}
        records.append(
            enrich_record(
                make_record(
                title=item.get("title", ""),
                abstract=abstract_from_openalex(item.get("abstract_inverted_index")),
                year=item.get("publication_year", ""),
                journal=source_info.get("display_name", ""),
                doi=doi_from_openalex(item.get("doi") or ""),
                url=item.get("id") or item.get("doi") or "",
                authors=authors,
                publication_type=item.get("type") or item.get("type_crossref") or "",
                source="openalex",
                network_group=network_group,
                network_term=network_term,
                search_query=query,
                citations=item.get("cited_by_count", ""),
                ),
                {
                    "sco_publisher": source_info.get("host_organization_name", ""),
                    "sco_pais": source_info.get("country_code", ""),
                    "sco_categorias": "; ".join(c.get("display_name", "") for c in item.get("topics") or [] if c.get("display_name")),
                },
            )
        )
    return records


def search_semantic_scholar(query: str, limit: int, network_group: str, network_term: str) -> list[dict[str, Any]]:
    fields = "title,year,abstract,venue,url,authors,externalIds,publicationTypes,citationCount"
    url = f"https://api.semanticscholar.org/graph/v1/paper/search?query={quote_plus(query)}&limit={limit}&fields={fields}"
    data = get_json(url)
    if not data:
        return []
    records = []
    for item in data.get("data") or []:
        external = item.get("externalIds") or {}
        authors = "; ".join(norm(a.get("name")) for a in item.get("authors") or [])
        records.append(
            make_record(
                title=item.get("title", ""),
                abstract=item.get("abstract", ""),
                year=item.get("year", ""),
                journal=item.get("venue", ""),
                doi=external.get("DOI", ""),
                url=item.get("url", ""),
                authors=authors,
                publication_type=", ".join(item.get("publicationTypes") or []),
                source="semantic_scholar",
                network_group=network_group,
                network_term=network_term,
                search_query=query,
                citations=item.get("citationCount", ""),
            )
        )
    return records


def search_crossref(query: str, limit: int, network_group: str, network_term: str) -> list[dict[str, Any]]:
    url = f"https://api.crossref.org/works?query={quote_plus(query)}&rows={limit}"
    data = get_json(url)
    if not data:
        return []
    records = []
    for item in (data.get("message") or {}).get("items") or []:
        title = (item.get("title") or [""])[0]
        published = item.get("published-print") or item.get("published-online") or item.get("created") or {}
        year = ((published.get("date-parts") or [[""]])[0] or [""])[0]
        authors = "; ".join(
            norm(f"{a.get('given', '')} {a.get('family', '')}") for a in item.get("author") or []
        )
        records.append(
            enrich_record(
                make_record(
                title=title,
                abstract=item.get("abstract", ""),
                year=year,
                journal=(item.get("container-title") or [""])[0],
                doi=item.get("DOI", ""),
                url=item.get("URL", ""),
                authors=authors,
                publication_type=item.get("type", ""),
                source="crossref",
                network_group=network_group,
                network_term=network_term,
                search_query=query,
                citations=item.get("is-referenced-by-count", ""),
                ),
                {
                    "volume": item.get("volume", "") or "x",
                    "numero": item.get("issue", "") or "x",
                    "paginas": item.get("page", "") or item.get("article-number", "") or "x",
                    "sco_publisher": item.get("publisher", ""),
                },
            )
        )
    return records


def search_arxiv(query: str, limit: int, network_group: str, network_term: str) -> list[dict[str, Any]]:
    url = f"https://export.arxiv.org/api/query?search_query=all:{quote_plus(query)}&max_results={limit}"
    try:
        response = SESSION.get(url, timeout=30)
        if response.status_code != 200:
            return []
        root = ET.fromstring(response.text)
    except Exception:
        return []
    ns = {"atom": "http://www.w3.org/2005/Atom", "arxiv": "http://arxiv.org/schemas/atom"}
    records = []
    for entry in root.findall("atom:entry", ns):
        authors = "; ".join(a.findtext("atom:name", default="", namespaces=ns) for a in entry.findall("atom:author", ns))
        published = entry.findtext("atom:published", default="", namespaces=ns)
        doi = entry.findtext("arxiv:doi", default="", namespaces=ns)
        records.append(
            make_record(
                title=entry.findtext("atom:title", default="", namespaces=ns),
                abstract=entry.findtext("atom:summary", default="", namespaces=ns),
                year=published[:4],
                journal="arXiv",
                doi=doi,
                url=entry.findtext("atom:id", default="", namespaces=ns),
                authors=authors,
                publication_type="preprint",
                source="arxiv",
                network_group=network_group,
                network_term=network_term,
                search_query=query,
            )
        )
    return records


def search_europe_pmc(query: str, limit: int, network_group: str, network_term: str) -> list[dict[str, Any]]:
    url = (
        "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
        f"?query={quote_plus(query)}&format=json&pageSize={limit}"
    )
    data = get_json(url)
    if not data:
        return []
    records = []
    for item in (data.get("resultList") or {}).get("result") or []:
        authors = item.get("authorString", "")
        records.append(
            make_record(
                title=item.get("title", ""),
                abstract=item.get("abstractText", ""),
                year=item.get("pubYear", ""),
                journal=item.get("journalTitle", ""),
                doi=item.get("doi", ""),
                url=item.get("fullTextUrlList", {}).get("fullTextUrl", [{}])[0].get("url", "") if item.get("fullTextUrlList") else "",
                authors=authors,
                publication_type=item.get("pubType", ""),
                source="europe_pmc",
                network_group=network_group,
                network_term=network_term,
                search_query=query,
                citations=item.get("citedByCount", ""),
            )
        )
    return records


def search_doaj(query: str, limit: int, network_group: str, network_term: str) -> list[dict[str, Any]]:
    url = f"https://doaj.org/api/search/articles/{quote_plus(query)}?pageSize={limit}"
    data = get_json(url)
    if not data:
        return []
    records = []
    for result in data.get("results") or []:
        bibjson = result.get("bibjson") or {}
        journal = bibjson.get("journal") or {}
        identifiers = bibjson.get("identifier") or []
        doi = ""
        for identifier in identifiers:
            if norm(identifier.get("type")).lower() == "doi":
                doi = identifier.get("id", "")
                break
        authors = "; ".join(norm(a.get("name")) for a in bibjson.get("author") or [])
        records.append(
            enrich_record(
                make_record(
                    title=bibjson.get("title", ""),
                    abstract=bibjson.get("abstract", ""),
                    year=(bibjson.get("year") or ""),
                    journal=journal.get("title", ""),
                    doi=doi,
                    url=(bibjson.get("link") or [{}])[0].get("url", ""),
                    authors=authors,
                    publication_type="journal-article",
                    source="doaj",
                    network_group=network_group,
                    network_term=network_term,
                    search_query=query,
                ),
                {
                    "volume": bibjson.get("volume", "") or "x",
                    "numero": bibjson.get("number", "") or "x",
                    "paginas": bibjson.get("start_page", "") or "x",
                    "sco_publisher": journal.get("publisher", ""),
                },
            )
        )
    return records


def enrich_record(record: dict[str, Any], values: dict[str, Any]) -> dict[str, Any]:
    for key, value in values.items():
        if key in record and not is_blank_for_merge(value):
            record[key] = value
    return record


def passes_year_filter(record: dict[str, Any], year_from: int | None, year_to: int | None) -> bool:
    try:
        year = int(record.get("ano") or 0)
    except ValueError:
        return True
    if year_from and year and year < year_from:
        return False
    if year_to and year and year > year_to:
        return False
    return True


def dedupe_key(record: dict[str, Any]) -> str:
    doi = norm(record.get("DOI")).lower()
    if doi:
        return f"doi:{doi}"
    return f"title:{normalize_key(record.get('titulo', ''))}"


def write_workbook(records: list[dict[str, Any]], output: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "G04_Henrique_Retail"
    ws.append(TEMPLATE_HEADERS)
    for record in records:
        ws.append([record.get(header, "") for header in TEMPLATE_HEADERS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 18,
        "B": 56,
        "C": 34,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 12,
        "I": 46,
        "J": 80,
    }
    for col in range(1, ws.max_column + 1):
        letter = ws.cell(1, col).column_letter
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[9].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    trace = wb.create_sheet("rastreamento_busca")
    trace_headers = ["DOI", "titulo", "fontes", "grupo_rede", "termo_rede", "consulta", "url", "citacoes", "tipo_publicacao"]
    trace.append(trace_headers)
    for record in records:
        trace.append(
            [
                record.get("DOI", ""),
                record.get("titulo", ""),
                record.get("_fonte_busca", ""),
                record.get("_grupo_rede", ""),
                record.get("_termo_rede", ""),
                record.get("_consulta", ""),
                record.get("_url", ""),
                record.get("_citacoes", ""),
                record.get("_tipo_publicacao", ""),
            ]
        )
    trace.freeze_panes = "A2"
    trace.auto_filter.ref = trace.dimensions
    for col in range(1, trace.max_column + 1):
        trace.column_dimensions[trace.cell(1, col).column_letter].width = 28
    for cell in trace[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def split_keywords(raw_keywords: list[str] | None) -> list[str]:
    if not raw_keywords:
        return []
    joined = " ".join(raw_keywords)
    parts = re.split(r"[\n;,]+", joined)
    return [norm(part.strip().strip('"')) for part in parts if norm(part.strip().strip('"'))]


def build_queries(args: argparse.Namespace) -> list[tuple[str, str, str]]:
    keywords = split_keywords(args.keywords)
    queries: list[tuple[str, str, str]] = []

    if args.mode in {"keywords", "both"} and keywords:
        for keyword in keywords:
            queries.append(("palavras_chave", keyword, keyword))

    if args.mode in {"networks", "both"}:
        for network_group, terms in NETWORK_TERMS.items():
            for term in terms:
                for context in args.context:
                    queries.append((network_group, term, f'"{context}" "{term}"'))

    if not queries:
        for network_group, terms in NETWORK_TERMS.items():
            for term in terms:
                for context in args.context:
                    queries.append((network_group, term, f'"{context}" "{term}"'))

    return queries


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Busca artigos sobre tipos de redes aplicados a retail/digital retail em varias fontes academicas.",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Planilha XLSX de saida.")
    parser.add_argument("--per-query", type=int, default=25, help="Resultados por consulta e fonte.")
    parser.add_argument("--wait", type=float, default=0.35, help="Pausa entre chamadas externas, em segundos.")
    parser.add_argument("--workers", type=int, default=8, help="Numero de buscas simultaneas.")
    parser.add_argument("--max-records", type=int, default=0, help="Limite final de artigos. 0 = sem limite.")
    parser.add_argument(
        "--keep-without-abstract",
        action="store_true",
        help="Mantem artigos sem abstract. Por padrao, eles sao descartados para acelerar e garantir planilha limpa.",
    )
    parser.add_argument(
        "--context",
        nargs="+",
        default=["retail", "digital retail", "omnichannel retail", "retail supply chain"],
        help="Contextos combinados com cada termo de rede.",
    )
    parser.add_argument(
        "--keywords",
        nargs="*",
        default=[],
        help="Palavras-chave/consultas livres. Separe por ponto e virgula, virgula ou nova linha.",
    )
    parser.add_argument(
        "--mode",
        choices=["networks", "keywords", "both"],
        default="networks",
        help="Modo de busca: redes predefinidas, palavras-chave livres ou ambos.",
    )
    parser.add_argument("--year-from", type=int, default=None, help="Ano inicial opcional.")
    parser.add_argument("--year-to", type=int, default=None, help="Ano final opcional.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sources = [search_openalex, search_semantic_scholar, search_crossref, search_europe_pmc, search_doaj, search_arxiv]
    records_by_key: dict[str, dict[str, Any]] = {}
    queries = build_queries(args)
    workers = max(1, args.workers)

    print(
        f"Consultas planejadas: {len(queries)} | fontes: {len(sources)} | trabalhadores: {workers}",
        flush=True,
    )
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = []
        for network_group, term, query in queries:
            print(f"Buscando: {query}", flush=True)
            for source in sources:
                futures.append(
                    executor.submit(
                        source,
                        query,
                        args.per_query,
                        network_group,
                        term,
                    )
                )

        completed = 0
        total = len(futures)
        for future in as_completed(futures):
            completed += 1
            try:
                source_records = future.result()
            except Exception as exc:
                print(f"[aviso] fonte falhou: {exc}", flush=True)
                continue
            for record in source_records:
                if not record.get("titulo"):
                    continue
                if not passes_year_filter(record, args.year_from, args.year_to):
                    continue
                key = dedupe_key(record)
                records_by_key[key] = merge_record(records_by_key.get(key), record)
            if completed == total or completed % max(1, workers) == 0:
                print(f"Progresso das fontes: {completed}/{total}", flush=True)
            if args.wait > 0:
                time.sleep(min(args.wait, 0.05))

    records = sorted(
        records_by_key.values(),
        key=lambda r: best_int(r.get("_citacoes", "")),
        reverse=True,
    )
    before_abstract_filter = len(records)
    if not args.keep_without_abstract:
        records = [record for record in records if norm(record.get("abstract"))]
    removed_without_abstract = before_abstract_filter - len(records)
    if args.max_records > 0:
        records = records[: args.max_records]
    write_workbook(records, args.output.resolve())
    print(f"Consultas executadas: {len(queries)}")
    print(f"Artigos sem abstract descartados: {removed_without_abstract}")
    print(f"Artigos unicos encontrados: {len(records)}")
    print(f"Planilha gerada: {args.output.resolve()}")


def best_int(value: Any) -> int:
    numbers = [int(n) for n in re.findall(r"\d+", norm(value))]
    return max(numbers) if numbers else 0


if __name__ == "__main__":
    main()
